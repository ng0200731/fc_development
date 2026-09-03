#!/usr/bin/env python3
"""
FC backend: SQLite database + REST API, also serves the static frontend.

Endpoints
---------
POST   /api/companies         body: {name, email_suffix}            -> create company
GET    /api/companies         -> [{id, name, email_suffix, created_at}]
GET    /api/companies/<id>    -> {id, name, email_suffix, members:[...]}
PUT    /api/companies/<id>    body: {name, email_suffix}            -> edit company
POST   /api/companies/<id>/members   body: {name, email_prefix, title, tel} -> add member
PUT    /api/members/<id>      body: {name, email_prefix, title, tel}-> edit member
DELETE /api/members/<id>      -> remove member
GET    /api/customers         -> flat list (company + its members) for the View page

Run:  python server.py   (default port 8088)
"""

import io
import json
import os
import re
import uuid
import sqlite3
import datetime
from http.server import HTTPServer, SimpleHTTPRequestHandler
from urllib.parse import urlparse, unquote, quote, parse_qs
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "data", "fc.db")
STATIC_DIR = os.path.join(BASE_DIR, "static")
UPLOADS_DIR = os.path.join(BASE_DIR, "data", "uploads")
PORT = int(os.environ.get("PORT", "8088"))

os.makedirs(UPLOADS_DIR, exist_ok=True)

# Directory scanned for sample images used by the Development / Create "Dummy"
# button and the image pool. Kept outside the project so the repo stays clean.
SAMPLE_IMAGES_DIR = r"C:\Users\ng\Desktop\canvas_source"
_sample_images_cache = None  # module-level cache for the scanned pool


def init_db():
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS companies (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            name          TEXT NOT NULL,
            email_suffix  TEXT NOT NULL,
            currency      TEXT,
            payment_term  TEXT,
            shipment_term TEXT,
            created_at    TEXT NOT NULL
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS members (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id    INTEGER NOT NULL REFERENCES companies(id) ON DELETE CASCADE,
            name          TEXT NOT NULL,
            email_prefix  TEXT NOT NULL,
            title         TEXT NOT NULL,
            tel           TEXT NOT NULL,
            created_at    TEXT NOT NULL
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS developments (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id    INTEGER,
            company_name  TEXT NOT NULL,
            member_id     INTEGER,
            member_name   TEXT,
            project_id    INTEGER,
            project_name  TEXT,
            item_name     TEXT NOT NULL,
            product_type  TEXT NOT NULL,
            height        REAL,
            width         REAL,
            raised_height REAL,
            no_of_color   INTEGER,
            pantones      TEXT,
            image_names   TEXT,
            doc_names     TEXT,
            material      TEXT,
            special       TEXT,
            remake        TEXT,
            created_at    TEXT NOT NULL,
            updated_at    TEXT NOT NULL
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS ship_to (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id  INTEGER NOT NULL REFERENCES companies(id) ON DELETE CASCADE,
            address     TEXT NOT NULL,
            is_default  INTEGER NOT NULL DEFAULT 0,
            created_at  TEXT NOT NULL
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS projects (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id  INTEGER NOT NULL REFERENCES companies(id) ON DELETE CASCADE,
            name        TEXT NOT NULL,
            created_at  TEXT NOT NULL,
            UNIQUE(company_id, name)
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS followups (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            development_id INTEGER NOT NULL REFERENCES developments(id) ON DELETE CASCADE,
            category       TEXT,
            note           TEXT,
            image_names    TEXT,
            doc_names      TEXT,
            created_at     TEXT NOT NULL
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS enquiries (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id    INTEGER,
            company_name  TEXT NOT NULL,
            member_id     INTEGER,
            member_name   TEXT,
            project_id    INTEGER,
            project_name  TEXT,
            item_name     TEXT,
            product_type  TEXT,
            height        REAL,
            width         REAL,
            raised_height REAL,
            no_of_color   INTEGER,
            pantones      TEXT,
            image_names   TEXT,
            doc_names     TEXT,
            material      TEXT,
            special       TEXT,
            remake        TEXT,
            color_sides   TEXT,
            notes         TEXT,
            created_at    TEXT NOT NULL,
            updated_at    TEXT NOT NULL
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS options (
            id       INTEGER PRIMARY KEY AUTOINCREMENT,
            level    TEXT NOT NULL,
            name     TEXT NOT NULL,
            value    TEXT NOT NULL,
            position INTEGER NOT NULL DEFAULT 0,
            UNIQUE(level, name, value)
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS product_type_factory (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            product_type TEXT NOT NULL,
            kind         TEXT NOT NULL,
            value        TEXT NOT NULL,
            position     INTEGER NOT NULL DEFAULT 0,
            UNIQUE(product_type, kind, value)
        )
        """
    )
    # Migrations: add any columns introduced after the initial schema so that
    # an existing database (already created without them) keeps working.
    _ensure_company_columns(conn)
    _ensure_dev_columns(conn)
    _ensure_enquiry_columns(conn)
    _ensure_option_groups_table(conn)
    _seed_option_groups(conn)
    _load_option_group_defs(conn)
    _migrate_options_table(conn)
    _seed_options(conn)
    _ensure_product_type_factory_table(conn)
    _seed_product_type_factory(conn)
    conn.commit()
    conn.close()


# Columns added to `companies` after the table was first created. Each entry
# is { column_name: SQL type }. init_db() adds any that are missing.
_COMPANY_MISSING_COLUMNS = {
    "currency": "TEXT",
    "payment_term": "TEXT",
    "shipment_term": "TEXT",
}


def _ensure_company_columns(conn):
    cur = conn.cursor()
    existing = {r[1] for r in cur.execute("PRAGMA table_info(companies)").fetchall()}
    for col, ctype in _COMPANY_MISSING_COLUMNS.items():
        if col not in existing:
            cur.execute(f"ALTER TABLE companies ADD COLUMN {col} {ctype}")


# Columns added to `developments` after the table was first created. Each entry
# is { column_name: SQL type }. init_db() adds any that are missing.
_DEV_MISSING_COLUMNS = {
    "doc_names": "TEXT",
    "project_id": "INTEGER",
    "project_name": "TEXT",
    "material": "TEXT",
    "special": "TEXT",
    "remake": "TEXT",
    "color_sides": "TEXT",
    "status": "TEXT",
}


def _ensure_dev_columns(conn):
    cur = conn.cursor()
    existing = {r[1] for r in cur.execute("PRAGMA table_info(developments)").fetchall()}
    for col, ctype in _DEV_MISSING_COLUMNS.items():
        if col not in existing:
            cur.execute(f"ALTER TABLE developments ADD COLUMN {col} {ctype}")
    # Existing rows predate the status column; give them the same initial state
    # as newly-created developments.
    if "status" in _DEV_MISSING_COLUMNS:
        cur.execute("UPDATE developments SET status = 'Created' WHERE status IS NULL OR TRIM(status) = ''")


# Columns added to `enquiries` after the table was first created. Each entry
# is { column_name: SQL type }. init_db() adds any that are missing.
_ENQUIRY_MISSING_COLUMNS = {
    "project_id": "INTEGER",
    "project_name": "TEXT",
    "product_type": "TEXT",
    "height": "REAL",
    "width": "REAL",
    "raised_height": "REAL",
    "no_of_color": "INTEGER",
    "pantones": "TEXT",
    "doc_names": "TEXT",
    "color_sides": "TEXT",
    # Free-text field captured on the Enquiry / Create screen ("Part 2 · Notes").
    # Round-trips through list / get / create / update.
    "notes": "TEXT",
}


def _ensure_enquiry_columns(conn):
    cur = conn.cursor()
    existing = {r[1] for r in cur.execute("PRAGMA table_info(enquiries)").fetchall()}
    for col, ctype in _ENQUIRY_MISSING_COLUMNS.items():
        if col not in existing:
            cur.execute(f"ALTER TABLE enquiries ADD COLUMN {col} {ctype}")


# Canonical SEED of every managed dropdown option set, grouped by menu level.
# `name` is the stable key used in the DB and the frontend; `label` is shown
# in the Settings UI. `seed` provides the initial values on a fresh database.
#
# At startup these are loaded into the live `_OPTION_GROUP_DEFS` dict AND
# persisted in the `option_groups` table, so new groups discovered by the
# Settings / Options "Refresh" scan and registered at runtime survive a server
# restart. The live dict is the single source of truth for validation/seed.
_OPTION_GROUP_SEED = {
    "customer": [
        {"name": "currency", "label": "Currency", "seed": ["USD", "RMB", "HKD"]},
        {"name": "payment_term", "label": "Payment term", "seed": ["COD", "credit 30 days", "credit 45 days"]},
        {"name": "shipment_term", "label": "Shipment term", "seed": ["Ex Work", "Door 2 Door", "FOB", "CIF"]},
    ],
    "development": [
        {"name": "product_type", "label": "Product type", "seed": [
            "woven label", "printed label", "screen print label", "hang tag",
            "raised silicon label", "heat transfer label", "leather patch", "embroidery patch",
        ]},
        {"name": "fabric", "label": "Fabric", "seed": ["polyester", "nylon", "cotton"]},
        {"name": "folding", "label": "Folding", "seed": [
            "loop fold", "end fold", "straight cut", "mitre fold", "Manhattan Fold", "Asymmetrical Fold",
        ]},
        {"name": "follow_up", "label": "Follow up", "seed": ["TBA"]},
    ],
    "enquiry": [
        # Enquiry / Create has no dropdowns (only Company & Member + Images).
        # Enquiry / Edit still shows a Product type select, but per the user's
        # instruction the managed sets must reflect the Create form, which has none.
    ],
}

# Live, mutable map of managed groups. Seeded from `_OPTION_GROUP_SEED` and the
# `option_groups` table at init; new groups registered via the API are appended
# here (and persisted in `option_groups`) so they survive restarts. Read it
# through `_option_group_defs()` rather than touching the dict directly.
_OPTION_GROUP_DEFS = {}


def _option_group_defs():
    return _OPTION_GROUP_DEFS


def _ensure_option_groups_table(conn):
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS option_groups (
            id     INTEGER PRIMARY KEY AUTOINCREMENT,
            level  TEXT NOT NULL,
            name   TEXT NOT NULL,
            label  TEXT NOT NULL DEFAULT '',
            UNIQUE(level, name)
        )
        """
    )


def _seed_option_groups(conn):
    cur = conn.cursor()
    # Add any newly-defined groups without disturbing groups/options already
    # present in an existing database.
    for level, groups in _OPTION_GROUP_SEED.items():
        for g in groups:
            cur.execute(
                "INSERT OR IGNORE INTO option_groups (level, name, label) VALUES (?, ?, ?)",
                (level, g["name"], g.get("label", g["name"])),
            )


def _load_option_group_defs(conn):
    global _OPTION_GROUP_DEFS
    # init_db()'s connection may not have row_factory set, so read by position.
    # Keep the canonical seed values attached to definitions because _seed_options
    # uses them when a newly-added group has no option rows yet.
    seed_by_key = {
        (level, g["name"]): g.get("seed", [])
        for level, groups in _OPTION_GROUP_SEED.items()
        for g in groups
    }
    defs = {lvl: [] for lvl in _OPTION_GROUP_SEED}
    for r in conn.execute("SELECT level, name, label FROM option_groups ORDER BY id").fetchall():
        level, name, label = r[0], r[1], r[2]
        defs.setdefault(level, []).append({
            "name": name,
            "label": label,
            "seed": seed_by_key.get((level, name), []),
        })
    _OPTION_GROUP_DEFS = defs


# Migrate the `options` table from the legacy shape (category, value, seq,
# owner_id) to the new (level, name, value, position) shape used by the
# Settings / Options manager. Runs once and is idempotent.
def _migrate_options_table(conn):
    cur = conn.cursor()
    cols = {r[1] for r in cur.execute("PRAGMA table_info(options)").fetchall()}
    # Legacy shape: has `category` but not our new `level`/`name` pair.
    if "category" in cols and "level" not in cols:
        # Legacy data stored product types under a single shared list
        # (category='product_type', one owner). Map it onto the two product_type
        # groups (development + enquiry) so the values survive the migration.
        cur.execute(
            "UPDATE options SET level='development', name='product_type' WHERE category='product_type'"
        )
        # Re-normalise seq -> position per (level, name) group.
        rows = cur.execute(
            "SELECT id, seq FROM options WHERE level IS NOT NULL"
        ).fetchall()
        for r in rows:
            cur.execute(
                "UPDATE options SET position=? WHERE id=?", (r[1] or 0, r[0])
            )
        # Rebuild without the legacy columns; add NOT NULL defaults for level/name.
        cur.execute("CREATE TABLE options_new ("
                    "id INTEGER PRIMARY KEY AUTOINCREMENT, "
                    "level TEXT NOT NULL DEFAULT '', name TEXT NOT NULL DEFAULT '', "
                    "value TEXT NOT NULL, position INTEGER NOT NULL DEFAULT 0, "
                    "UNIQUE(level, name, value))")
        cur.execute("INSERT INTO options_new (id, level, name, value, position) "
                    "SELECT id, COALESCE(level, ''), COALESCE(name, ''), value, position FROM options")
        cur.execute("DROP TABLE options")
        cur.execute("ALTER TABLE options_new RENAME TO options")
        cols = {r[1] for r in cur.execute("PRAGMA table_info(options)").fetchall()}
    # Already in the new shape, or just migrated — ensure level/name are present.
    if "level" not in cols or "name" not in cols:
        cur.execute("ALTER TABLE options ADD COLUMN level TEXT NOT NULL DEFAULT ''")
        cur.execute("ALTER TABLE options ADD COLUMN name TEXT NOT NULL DEFAULT ''")
    if "position" not in cols:
        cur.execute("ALTER TABLE options ADD COLUMN position INTEGER NOT NULL DEFAULT 0")


# Seed the `options` table once, on a fresh database (empty table), so the
# dropdown lists start with today's hardcoded values. Existing rows are kept.
def _register_option_group(conn, level, name, label):
    """Persist a newly-discovered managed group so it survives restarts and is
    visible to the Settings UI. No-op if the (level, name) group already exists.
    `label` falls back to the provided name when omitted."""
    cur = conn.cursor()
    exists = cur.execute(
        "SELECT 1 FROM option_groups WHERE level=? AND name=?", (level, name)
    ).fetchone()
    if exists:
        return False
    cur.execute(
        "INSERT INTO option_groups (level, name, label) VALUES (?, ?, ?)",
        (level, name, label or name),
    )
    # Surface it in the live dict immediately so subsequent requests see it.
    _OPTION_GROUP_DEFS.setdefault(level, []).append({"name": name, "label": label or name})
    return True


def _seed_options(conn):
    cur = conn.cursor()
    # Seed per (level, name) group so that only genuinely empty groups get the
    # default values — existing data (including migrated legacy rows) is kept.
    for level, groups in _OPTION_GROUP_DEFS.items():
        for g in groups:
            have = cur.execute(
                "SELECT 1 FROM options WHERE level=? AND name=?", (level, g["name"])
            ).fetchone()
            if have:
                continue
            for p, v in enumerate(g["seed"]):
                cur.execute(
                    "INSERT INTO options (level, name, value, position) VALUES (?, ?, ?, ?)",
                    (level, g["name"], v, p),
                )


def _options_payload():
    """Build the full {level: [{name,label,values:[{id,value}]}]} payload."""
    conn = db()
    rows = conn.execute(
        "SELECT id, level, name, value, position FROM options ORDER BY level, name, position"
    ).fetchall()
    conn.close()
    by_key = {}
    for r in rows:
        by_key.setdefault((r["level"], r["name"]), []).append({"id": r["id"], "value": r["value"]})
    out = {}
    for level, groups in _OPTION_GROUP_DEFS.items():
        out[level] = [
            {"name": g["name"], "label": g["label"], "values": by_key.get((level, g["name"]), [])}
            for g in groups
        ]
    return out


def api_get_options(handler):
    return json_response(handler, _options_payload())


def api_add_option(handler):
    data = read_json_body(handler)
    level = (data.get("level") or "").strip()
    name = (data.get("name") or "").strip()
    value = (data.get("value") or "").strip()
    if level not in _OPTION_GROUP_DEFS or not name or not value:
        return json_response(handler, {"error": "level, name and value are required"}, 400)
    conn = db()
    exists = conn.execute(
        "SELECT 1 FROM options WHERE level=? AND name=? AND value=?", (level, name, value)
    ).fetchone()
    if exists:
        conn.close()
        return json_response(handler, {"error": "option already exists"}, 409)
    max_pos = conn.execute(
        "SELECT COALESCE(MAX(position), -1) AS m FROM options WHERE level=? AND name=?", (level, name)
    ).fetchone()["m"]
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO options (level, name, value, position) VALUES (?, ?, ?, ?)",
        (level, name, value, max_pos + 1),
    )
    oid = cur.lastrowid
    row = conn.execute(
        "SELECT id, level, name, value, position FROM options WHERE id=?", (oid,)
    ).fetchone()
    conn.commit()
    conn.close()
    return json_response(handler, dict(row), 201)


def api_reorder_options(handler):
    data = read_json_body(handler)
    level = (data.get("level") or "").strip()
    name = (data.get("name") or "").strip()
    ordered = data.get("orderedValues")
    if level not in _OPTION_GROUP_DEFS or not name or not isinstance(ordered, list):
        return json_response(handler, {"error": "level, name and orderedValues[] required"}, 400)
    conn = db()
    # Only move values that actually belong to this (level, name) group.
    valid = {r["value"] for r in conn.execute(
        "SELECT value FROM options WHERE level=? AND name=?", (level, name)).fetchall()}
    pos = 0
    for v in ordered:
        if v in valid:
            conn.execute(
                "UPDATE options SET position=? WHERE level=? AND name=? AND value=?",
                (pos, level, name, v),
            )
            pos += 1
    conn.commit()
    conn.close()
    return json_response(handler, {"ok": True}, 200)


def api_rename_option(handler, oid):
    conn = db()
    row = conn.execute("SELECT * FROM options WHERE id=?", (oid,)).fetchone()
    if not row:
        conn.close()
        return json_response(handler, {"error": "not found"}, 404)
    data = read_json_body(handler)
    value = (data.get("value") or "").strip()
    if not value:
        conn.close()
        return json_response(handler, {"error": "value is required"}, 400)
    clash = conn.execute(
        "SELECT 1 FROM options WHERE level=? AND name=? AND value=? AND id!=?",
        (row["level"], row["name"], value, oid),
    ).fetchone()
    if clash:
        conn.close()
        return json_response(handler, {"error": "option already exists"}, 409)
    conn.execute("UPDATE options SET value=? WHERE id=?", (value, oid))
    conn.commit()
    conn.close()
    return json_response(handler, {"ok": True, "id": oid}, 200)


def api_delete_option(handler, oid):
    conn = db()
    row = conn.execute("SELECT * FROM options WHERE id=?", (oid,)).fetchone()
    if not row:
        conn.close()
        return json_response(handler, {"error": "not found"}, 404)
    conn.execute("DELETE FROM options WHERE id=?", (oid,))
    # Re-pack positions so the remaining list keeps a clean 0..n-1 order.
    remaining = conn.execute(
        "SELECT id FROM options WHERE level=? AND name=? ORDER BY position",
        (row["level"], row["name"]),
    ).fetchall()
    for i, r in enumerate(remaining):
        conn.execute("UPDATE options SET position=? WHERE id=?", (i, r["id"]))
    conn.commit()
    conn.close()
    return json_response(handler, {"ok": True, "id": oid}, 200)


# Return every currently-managed (level, name) group as a flat list of
# {level, name, label}. Used by the Settings / Options "Refresh" scan to diff
# against the dropdowns the forms actually render.
def api_list_groups(handler):
    conn = db()
    rows = conn.execute(
        "SELECT level, name, label FROM option_groups ORDER BY level, id"
    ).fetchall()
    conn.close()
    return json_response(handler, [
        {"level": r["level"], "name": r["name"], "label": r["label"]} for r in rows
    ])


# Register a newly-discovered dropdown group. The scan in the Settings UI calls
# this after it finds a <select> whose (level, name) is not yet managed. The
# group is persisted so it survives a restart and shows in the Settings dropdowns.
def api_register_group(handler):
    data = read_json_body(handler)
    level = (data.get("level") or "").strip()
    name = (data.get("name") or "").strip()
    label = (data.get("label") or "").strip()
    if not level or not name:
        return json_response(handler, {"error": "level and name are required"}, 400)
    conn = db()
    is_new = _register_option_group(conn, level, name, label)
    conn.commit()
    conn.close()
    return json_response(handler, {"level": level, "name": name, "label": label, "isNew": is_new}, 200)


# ---------------------------------------------------------------------------
# Product type "factory": per (product_type, kind) overrides for the Fabric and
# Folding dropdowns. Product types not present here fall back to the global
# Development -> Fabric / Folding lists. Persisted so edits survive restarts.
#
# GENERALIZED MODEL: each product type owns an ordered set of *named lists*
# (the `kind` column). "fabric" and "folding" are the two default list names
# but any name is allowed (e.g. "Material", "Finish", "GSM") — see Part 4 of
# the spec. `list_seq` orders the lists *within* a product type; `position`
# orders the options *within* a single list.
# ---------------------------------------------------------------------------

def _ensure_product_type_factory_table(conn):
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS product_type_factory (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            product_type TEXT NOT NULL,
            kind         TEXT NOT NULL,
            value        TEXT NOT NULL,
            position     INTEGER NOT NULL DEFAULT 0,
            list_seq     INTEGER NOT NULL DEFAULT 0,
            input_type   TEXT NOT NULL DEFAULT 'dropdown',
            UNIQUE(product_type, kind, value)
        )
        """
    )
    # Migrations for older databases (created before these columns existed).
    # Add the missing column in place so an existing DB keeps working without a
    # rebuild. `input_type` defaults to 'dropdown' so legacy rows behave as a
    # plain dropdown — matching the previous rendering.
    existing = {r[1] for r in conn.execute("PRAGMA table_info(product_type_factory)").fetchall()}
    if "list_seq" not in existing:
        conn.execute("ALTER TABLE product_type_factory ADD COLUMN list_seq INTEGER NOT NULL DEFAULT 0")
    if "input_type" not in existing:
        conn.execute(
            "ALTER TABLE product_type_factory ADD COLUMN input_type TEXT NOT NULL DEFAULT 'dropdown'"
        )


# Valid values for the per-kind input_type column. Anything not in this set is
# coerced back to 'dropdown' on write so the database never stores garbage.
_PTF_INPUT_TYPES = {"dropdown", "radio", "text", "textarea"}


def _coerce_input_type(value):
    """Normalize an incoming input_type; falls back to 'dropdown' on garbage."""
    v = (value or "").strip().lower()
    return v if v in _PTF_INPUT_TYPES else "dropdown"


def _set_ptf_list_input_type(conn, product_type, kind, input_type):
    """Write a single input_type value to every row of one (product_type, kind)."""
    conn.execute(
        "UPDATE product_type_factory SET input_type=? WHERE product_type=? AND kind=?",
        (_coerce_input_type(input_type), product_type, kind),
    )


# Seed the factory with sensible defaults (idempotent, per product type).
#
# Per the user's request, NO product type is pre-seeded with Fabric / Folding
# (or any other list). Every product type starts empty in the Product Type
# Factory panel so the admin picks exactly which named lists each product
# needs (Material, Finish, GSM, …). Product types with no override fall back
# to the global Development -> Fabric / Folding lists via the frontend's
# `fabricOptionsFor` / `foldingOptionsFor` helpers, so this change has no
# runtime impact — only the Settings UI shows a blank slate per product.
#
# The function body is kept as a no-op so the init_db() call site stays the
# same and any future re-seeding (e.g. an admin "Reset to defaults" button)
# has an obvious place to live. Existing rows are NEVER touched — once a user
# adds a list for a product, deletion sticks across restarts.
def _seed_product_type_factory(conn):
    # Intentionally empty: see comment above.
    return


def _next_list_seq(conn, product_type):
    row = conn.execute(
        "SELECT COALESCE(MAX(list_seq), -1) AS m FROM product_type_factory WHERE product_type=?",
        (product_type,),
    ).fetchone()
    return row["m"] + 1


def api_get_product_type_factory(handler):
    conn = db()
    rows = conn.execute(
        "SELECT id, product_type, kind, value, position, list_seq, input_type "
        "FROM product_type_factory "
        "ORDER BY product_type, list_seq, position"
    ).fetchall()
    conn.close()
    out = {}
    # Pre-seed every known product type as an empty object so the frontend can
    # tell "this product type exists but has no lists yet" (initialized, empty)
    # apart from "this product type was never loaded" (uninitialized). This
    # keeps a deleted list from resurrecting phantom Fabric/Folding blocks.
    for pt in (
        "woven label", "printed label", "screen print label", "hang tag",
        "raised silicon label", "heat transfer label", "leather patch", "embroidery patch",
    ):
        out.setdefault(pt, {})
    for r in rows:
        prod = out.setdefault(r["product_type"], {})
        prod.setdefault(r["kind"], []).append({
            "id": r["id"], "value": r["value"], "position": r["position"],
            "input_type": r["input_type"],
        })
    return json_response(handler, out)


def api_add_ptf(handler):
    data = read_json_body(handler)
    product_type = (data.get("product_type") or "").strip()
    kind = (data.get("kind") or "").strip()
    value = (data.get("value") or "").strip()
    if not product_type or not kind or not value:
        return json_response(handler, {"error": "product_type, kind and value are required"}, 400)
    conn = db()
    exists = conn.execute(
        "SELECT 1 FROM product_type_factory WHERE product_type=? AND kind=? AND value=?",
        (product_type, kind, value),
    ).fetchone()
    if exists:
        conn.close()
        return json_response(handler, {"error": "value already exists"}, 409)
    max_pos = conn.execute(
        "SELECT COALESCE(MAX(position), -1) AS m FROM product_type_factory WHERE product_type=? AND kind=?",
        (product_type, kind),
    ).fetchone()["m"]
    cur = conn.cursor()
    # A brand-new list (no rows yet) gets the next list_seq so it renders last.
    has_any = cur.execute(
        "SELECT 1 FROM product_type_factory WHERE product_type=? AND kind=?",
        (product_type, kind),
    ).fetchone()
    list_seq = 0 if has_any else _next_list_seq(conn, product_type)
    cur.execute(
        "INSERT INTO product_type_factory (product_type, kind, value, position, list_seq) "
        "VALUES (?, ?, ?, ?, ?)",
        (product_type, kind, value, max_pos + 1, list_seq),
    )
    oid = cur.lastrowid
    row = conn.execute(
        "SELECT id, product_type, kind, value, position, list_seq FROM product_type_factory WHERE id=?", (oid,)
    ).fetchone()
    conn.commit()
    conn.close()
    return json_response(handler, dict(row), 201)


def api_reorder_ptf(handler):
    data = read_json_body(handler)
    product_type = (data.get("product_type") or "").strip()
    kind = (data.get("kind") or "").strip()
    ordered = data.get("orderedValues")
    if not product_type or not kind or not isinstance(ordered, list):
        return json_response(handler, {"error": "product_type, kind and orderedValues[] required"}, 400)
    conn = db()
    valid = {r["value"] for r in conn.execute(
        "SELECT value FROM product_type_factory WHERE product_type=? AND kind=?",
        (product_type, kind)).fetchall()}
    pos = 0
    for v in ordered:
        if v in valid:
            conn.execute(
                "UPDATE product_type_factory SET position=? WHERE product_type=? AND kind=? AND value=?",
                (pos, product_type, kind, v),
            )
            pos += 1
    conn.commit()
    conn.close()
    return json_response(handler, {"ok": True}, 200)


# --- List-level operations (create / rename / delete / reorder a whole list) --

def api_add_ptf_list(handler):
    """Create an empty list (no options yet) for a product type."""
    data = read_json_body(handler)
    product_type = (data.get("product_type") or "").strip()
    kind = (data.get("kind") or "").strip()
    input_type = _coerce_input_type(data.get("input_type"))
    if not product_type or not kind:
        return json_response(handler, {"error": "product_type and kind are required"}, 400)
    conn = db()
    clash = conn.execute(
        "SELECT 1 FROM product_type_factory WHERE product_type=? AND kind=?",
        (product_type, kind),
    ).fetchone()
    if clash:
        conn.close()
        return json_response(handler, {"error": "list already exists"}, 409)
    list_seq = _next_list_seq(conn, product_type)
    conn.execute(
        "INSERT INTO product_type_factory (product_type, kind, value, position, list_seq, input_type) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        (product_type, kind, "__placeholder__", 0, list_seq, input_type),
    )
    oid = conn.execute(
        "SELECT id FROM product_type_factory WHERE product_type=? AND kind=? AND value=?",
        (product_type, kind, "__placeholder__"),
    ).fetchone()["id"]
    conn.commit()
    conn.close()
    return json_response(handler, {"ok": True, "product_type": product_type, "kind": kind, "input_type": input_type}, 201)


def api_set_ptf_list_input_type(handler):
    """Set the input_type of an existing list (every row for (product_type, kind))."""
    data = read_json_body(handler)
    product_type = (data.get("product_type") or "").strip()
    kind = (data.get("kind") or "").strip()
    input_type = _coerce_input_type(data.get("input_type"))
    if not product_type or not kind:
        return json_response(handler, {"error": "product_type and kind are required"}, 400)
    conn = db()
    exists = conn.execute(
        "SELECT 1 FROM product_type_factory WHERE product_type=? AND kind=?",
        (product_type, kind),
    ).fetchone()
    if not exists:
        conn.close()
        return json_response(handler, {"error": "list not found"}, 404)
    _set_ptf_list_input_type(conn, product_type, kind, input_type)
    conn.commit()
    conn.close()
    return json_response(handler, {"ok": True, "product_type": product_type, "kind": kind, "input_type": input_type}, 200)


def api_rename_ptf_list(handler):
    """Rename a whole list (every option row for (product_type, oldKind))."""
    data = read_json_body(handler)
    product_type = (data.get("product_type") or "").strip()
    old_kind = (data.get("oldKind") or "").strip()
    new_kind = (data.get("newKind") or "").strip()
    if not product_type or not old_kind or not new_kind:
        return json_response(handler, {"error": "product_type, oldKind and newKind are required"}, 400)
    if old_kind == new_kind:
        return json_response(handler, {"ok": True})
    conn = db()
    existing = conn.execute(
        "SELECT 1 FROM product_type_factory WHERE product_type=? AND kind=?",
        (product_type, new_kind),
    ).fetchone()
    if existing:
        conn.close()
        return json_response(handler, {"error": "a list named '%s' already exists" % new_kind}, 409)
    source = conn.execute(
        "SELECT 1 FROM product_type_factory WHERE product_type=? AND kind=?",
        (product_type, old_kind),
    ).fetchone()
    if not source:
        conn.close()
        return json_response(handler, {"error": "list not found"}, 404)
    conn.execute(
        "UPDATE product_type_factory SET kind=? WHERE product_type=? AND kind=?",
        (new_kind, product_type, old_kind),
    )
    conn.commit()
    conn.close()
    return json_response(handler, {"ok": True, "product_type": product_type, "kind": new_kind}, 200)


def api_delete_ptf_list(handler):
    """Delete an entire list (all option rows) for a product type."""
    data = read_json_body(handler)
    product_type = (data.get("product_type") or "").strip()
    kind = (data.get("kind") or "").strip()
    if not product_type or not kind:
        return json_response(handler, {"error": "product_type and kind are required"}, 400)
    conn = db()
    exists = conn.execute(
        "SELECT 1 FROM product_type_factory WHERE product_type=? AND kind=?",
        (product_type, kind),
    ).fetchone()
    if not exists:
        conn.close()
        return json_response(handler, {"error": "list not found"}, 404)
    conn.execute(
        "DELETE FROM product_type_factory WHERE product_type=? AND kind=?",
        (product_type, kind),
    )
    # Re-pack list_seq so the remaining lists stay contiguous / ordered.
    remaining = conn.execute(
        "SELECT DISTINCT kind FROM product_type_factory WHERE product_type=? ORDER BY list_seq",
        (product_type,),
    ).fetchall()
    for i, r in enumerate(remaining):
        conn.execute(
            "UPDATE product_type_factory SET list_seq=? WHERE product_type=? AND kind=?",
            (i, product_type, r["kind"]),
        )
    conn.commit()
    conn.close()
    return json_response(handler, {"ok": True, "product_type": product_type, "kind": kind}, 200)


def api_reorder_ptf_lists(handler):
    """Reorder the lists of a product type. Body: {product_type, orderedKinds:[...]}."""
    data = read_json_body(handler)
    product_type = (data.get("product_type") or "").strip()
    ordered = data.get("orderedKinds")
    if not product_type or not isinstance(ordered, list):
        return json_response(handler, {"error": "product_type and orderedKinds[] required"}, 400)
    conn = db()
    valid = {r["kind"] for r in conn.execute(
        "SELECT DISTINCT kind FROM product_type_factory WHERE product_type=?",
        (product_type,)).fetchall()}
    seq = 0
    for k in ordered:
        if k in valid:
            conn.execute(
                "UPDATE product_type_factory SET list_seq=? WHERE product_type=? AND kind=?",
                (seq, product_type, k),
            )
            seq += 1
    conn.commit()
    conn.close()
    return json_response(handler, {"ok": True}, 200)


def api_rename_ptf(handler, oid):
    conn = db()
    row = conn.execute("SELECT * FROM product_type_factory WHERE id=?", (oid,)).fetchone()
    if not row:
        conn.close()
        return json_response(handler, {"error": "not found"}, 404)
    data = read_json_body(handler)
    value = (data.get("value") or "").strip()
    if not value:
        conn.close()
        return json_response(handler, {"error": "value is required"}, 400)
    clash = conn.execute(
        "SELECT 1 FROM product_type_factory WHERE product_type=? AND kind=? AND value=? AND id!=?",
        (row["product_type"], row["kind"], value, oid),
    ).fetchone()
    if clash:
        conn.close()
        return json_response(handler, {"error": "value already exists"}, 409)
    conn.execute("UPDATE product_type_factory SET value=? WHERE id=?", (value, oid))
    conn.commit()
    conn.close()
    return json_response(handler, {"ok": True, "id": oid}, 200)


def api_delete_ptf(handler, oid):
    conn = db()
    row = conn.execute("SELECT * FROM product_type_factory WHERE id=?", (oid,)).fetchone()
    if not row:
        conn.close()
        return json_response(handler, {"error": "not found"}, 404)
    conn.execute("DELETE FROM product_type_factory WHERE id=?", (oid,))
    remaining = conn.execute(
        "SELECT id FROM product_type_factory WHERE product_type=? AND kind=? ORDER BY position",
        (row["product_type"], row["kind"]),
    ).fetchall()
    for i, r in enumerate(remaining):
        conn.execute("UPDATE product_type_factory SET position=? WHERE id=?", (i, r["id"]))
    # Keep the list alive even when its last option is removed (so an empty list
    # still shows up in the UI). Re-insert a placeholder if the list is now empty.
    still = conn.execute(
        "SELECT 1 FROM product_type_factory WHERE product_type=? AND kind=?",
        (row["product_type"], row["kind"]),
    ).fetchone()
    if not still:
        list_seq = conn.execute(
            "SELECT COALESCE(MIN(list_seq),0) FROM product_type_factory WHERE product_type=?",
            (row["product_type"],),
        ).fetchone()[0]
        conn.execute(
            "INSERT INTO product_type_factory (product_type, kind, value, position, list_seq) "
            "VALUES (?, ?, ?, ?, ?)",
            (row["product_type"], row["kind"], "__placeholder__", 0, list_seq),
        )
    conn.commit()
    conn.close()
    return json_response(handler, {"ok": True, "id": oid}, 200)


def now_iso():
    return datetime.datetime.now().isoformat(timespec="seconds")


def json_response(handler, payload, status=200):
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Content-Length", str(len(body)))
    handler.send_header("Access-Control-Allow-Origin", "*")
    handler.send_header("Access-Control-Allow-Methods", "GET, POST, PUT, DELETE, OPTIONS")
    handler.send_header("Access-Control-Allow-Headers", "Content-Type")
    handler.end_headers()
    handler.wfile.write(body)


def read_json_body(handler):
    length = int(handler.headers.get("Content-Length", 0) or 0)
    if length == 0:
        return {}
    raw = handler.rfile.read(length)
    return json.loads(raw.decode("utf-8"))


def db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


# --- Excel export (Development / Enquiry / Customer views) -------------------
#
# Each view exports a real .xlsx (via openpyxl) with one row per record and a
# thumbnail image embedded in an "Image" column. Image bytes are read straight
# off disk (sample images under SAMPLE_IMAGES_DIR, uploads under UPLOADS_DIR)
# so the export works offline — no round-trip through the browser. Images that
# can't be resolved are left blank rather than failing the whole export.

IMG_THUMB_W = 110  # px, exported image width (height scaled to keep ratio)
IMG_THUMB_H = 80   # px, exported image height (used for row height + grid)


def display_name(name):
    """Strip the embedded '<uuid>__' prefix (and any path) so uploaded files
    show their original filename; sample paths show just the basename."""
    if not name:
        return ""
    base = name.split("/")[-1] if "/" in name else name
    return re.sub(r"^[^_]*__", "", base)


def _resolve_image_path(name):
    """Map a stored image name to an on-disk absolute path, or None.

    - "uploads/<uuid>__<orig>" or bare "uploads/..." -> UPLOADS_DIR (best-effort)
    - anything else (legacy sample / Dummy)          -> SAMPLE_IMAGES_DIR
    """
    if not name:
        return None
    if name.startswith("uploads/"):
        rel = name[len("uploads/"):]
        p = safe_upload_path(rel)
        if p:
            return p
        resolved = resolve_upload_name(rel)
        if resolved:
            return os.path.join(UPLOADS_DIR, resolved)
        return None
    rel = name
    p = safe_sample_path(rel)
    if p:
        return p
    # Try the bare basename under the sample root (legacy names).
    cand = os.path.join(SAMPLE_IMAGES_DIR, os.path.basename(rel))
    return cand if os.path.isfile(cand) else None


def _embed_thumbnail(ws, cell, name):
    """Embed an image (scaled to a thumbnail) into `cell`. Returns True on success."""
    path = _resolve_image_path(name)
    if not path:
        return False
    try:
        img = XLImage(path)
    except Exception:
        return False
    # Scale uniformly so the larger dimension fits the thumbnail box.
    w, h = img.width, img.height
    if w <= 0 or h <= 0:
        return False
    scale = min(IMG_THUMB_W / w, IMG_THUMB_H / h, 1.0)
    img.width = int(w * scale)
    img.height = int(h * scale)
    ws.add_image(img, cell)
    return True


def _style_header(ws, ncols):
    fill = PatternFill("solid", fgColor="1F2937")
    font = Font(bold=True, color="FFFFFF")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center", horizontal="left")


# Thin all-side border applied to every cell of an exported sheet.
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


def _apply_borders(ws, nrows, ncols):
    """Draw a thin border around every cell in the used range."""
    for row in range(1, nrows + 1):
        for col in range(1, ncols + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER


def _build_workbook(records, sheet_title):
    """Build an openpyxl Workbook from `records`.

    Each record is (headers, cells, image_names):
      - headers:    list of column titles (same for every record)
      - cells:      list of text values, one PER COLUMN (should match headers)
      - image_names: stored image names; the FIRST is embedded as a thumbnail
                     into whichever column is titled exactly "Image"

    Written so each header column gets exactly one cell (short records are
    padded with ""), so values can never drift into the wrong column.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    if not records:
        ws.cell(row=1, column=1, value="(no data)")
        return wb

    headers = list(records[0][0])
    ncols = len(headers)

    # Header row: dark fill, bold white text, thin border.
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = PatternFill("solid", fgColor="1F2937")
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(vertical="center", horizontal="left")
        c.border = THIN_BORDER

    # Find the "Image" column (1-based).
    img_col = headers.index("Image") + 1 if "Image" in headers else None

    # One data row per record.
    for ri, (_headers, cells, image_names) in enumerate(records, start=2):
        for ci in range(1, ncols + 1):
            # One value per column — pad with "" if the record is short.
            val = cells[ci - 1] if ci - 1 < len(cells) else ""
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
            c.border = THIN_BORDER
        if img_col is not None and image_names:
            ws.row_dimensions[ri].height = IMG_THUMB_H + 6
            _embed_thumbnail(ws, f"{get_column_letter(img_col)}{ri}", image_names[0])

    # Lock the Image column width to the thumbnail so the picture fits the cell.
    if img_col is not None:
        ws.column_dimensions[get_column_letter(img_col)].width = (IMG_THUMB_W + 6) / 7.0

    return wb


def _send_xlsx(handler, wb, filename):
    out = io.BytesIO()
    wb.save(out)
    data = out.getvalue()
    handler.send_response(200)
    handler.send_header(
        "Content-Type",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    handler.send_header("Content-Length", str(len(data)))
    ascii_name = filename.encode("ascii", "ignore").decode("ascii") or "export"
    handler.send_header(
        "Content-Disposition",
        "attachment; filename=\"%s\"; filename*=UTF-8''%s"
        % (ascii_name, quote(filename)),
    )
    handler.send_header("Cache-Control", "no-store")
    handler.end_headers()
    handler.wfile.write(data)


def _dev_record_to_xlsx(d):
    """Flat (headers, cells, image_names) for one development/enquiry row.

    Column order mirrors the Development / View grid exactly:
    Company, Member, Item, Product Type, Image, Documents, Material, Special,
    Height (mm), Width (mm), Remark, Created, Updated, Details.
    `cells` has exactly one entry per header, in order.
    """
    images = d.get("image_names") or []
    docs = d.get("doc_names") or []
    headers = ["Company", "Member", "Item", "Product Type", "Image",
               "Documents", "Material", "Special",
               "Height (mm)", "Width (mm)", "Remark",
               "Created", "Updated", "Details"]
    material = d.get("material")
    special = d.get("special")
    height = d.get("height")
    width = d.get("width")
    # `remake` may already be a list (JSON-decoded) or a JSON string; normalize.
    remake = d.get("remake")
    if isinstance(remake, str):
        try:
            import json as _json
            remake = _json.loads(remake)
        except Exception:
            remake = [remake]
    if not isinstance(remake, list):
        remake = []
    cells = [
        d.get("company_name") or "",
        d.get("member_name") or "",
        d.get("item_name") or "",
        d.get("product_type") or "",
        "",                              # Image column: holds the embedded thumbnail
        _docs_summary(docs),
        _material_summary(material),
        _special_summary(special),
        "" if height is None else str(height),
        "" if width is None else str(width),
        "\n".join(remake),
        d.get("created_at") or "",
        d.get("updated_at") or "",
        dev_details_summary(d),          # Details: size / colors summary
    ]
    return headers, cells, images


def _material_summary(mat):
    """Match the JS `materialSummary()`: short text for the Material column."""
    if not mat or not isinstance(mat, dict):
        return ""
    parts = []
    if mat.get("recycle"):
        parts.append("recycle" if mat["recycle"] == "recycle" else "non-recycle")
    if mat.get("fabric"):
        parts.append(mat["fabric"])
    if mat.get("edge"):
        parts.append("slit edge" if mat["edge"] == "slit" else "woven edge")
    if mat.get("folding"):
        parts.append(mat["folding"])
    lists = mat.get("lists")
    if lists and isinstance(lists, dict):
        for k in lists.keys():
            if lists[k]:
                parts.append(f"{k}: {lists[k]}")
    return " · ".join(parts)


def _special_summary(spec):
    """Match the JS `specialSummary()`: short text for the Special column."""
    if not spec or not isinstance(spec, dict):
        return ""
    if spec.get("variable"):
        return "variable" if spec["variable"] == "variable" else "non variable"
    return ""


def _docs_summary(doc_names):
    """Render the Documents column like the View grid: a '📄 name' line each."""
    docs = doc_names or []
    if not docs:
        return "—"
    return "\n".join("📄 " + display_name(n) for n in docs)


def _customer_record_to_xlsx(c):
    members = c.get("members") or []
    ship_to = " | ".join(s.get("address", "") for s in (c.get("ship_to") or []))
    projects = ", ".join(p.get("name", "") for p in (c.get("projects") or []))
    rows = []
    if not members:
        members = [None]
    for m in members:
        headers = ["Company Name", "Member Name", "Member Email", "Title",
                   "Tel", "Currency", "Payment Term", "Shipment Term",
                   "Ship To", "Projects"]
        cells = [
            c.get("name") or "",
            m.get("name") if m else "",
            (m.get("email_prefix") + "@" + c.get("email_suffix")) if m else "",
            m.get("title") if m else "",
            m.get("tel") if m else "",
            c.get("currency") or "",
            c.get("payment_term") or "",
            c.get("shipment_term") or "",
            ship_to,
            projects,
        ]
        rows.append((headers, cells, []))
    return rows


def api_export_developments(handler):
    conn = db()
    # `?ids=1,2,3` filters the export to only the checked rows from the View grid.
    qs = parse_qs(urlparse(handler.path).query)
    ids_raw = (qs.get("ids") or [""])[0]
    if ids_raw:
        try:
            ids = [int(x) for x in ids_raw.split(",") if x.strip()]
        except ValueError:
            ids = []
        if not ids:
            rows = []
        else:
            placeholders = ",".join("?" * len(ids))
            rows = conn.execute(
                f"SELECT * FROM developments WHERE id IN ({placeholders}) ORDER BY id DESC",
                ids,
            ).fetchall()
    else:
        rows = conn.execute("SELECT * FROM developments ORDER BY id DESC").fetchall()
    conn.close()
    records = [_dev_record_to_xlsx(_dev_row_to_payload(r)) for r in rows]
    wb = _build_workbook(records, "Developments")
    _send_xlsx(handler, wb, "developments.xlsx")


def api_export_enquiries(handler):
    conn = db()
    rows = conn.execute("SELECT * FROM enquiries ORDER BY id DESC").fetchall()
    conn.close()
    records = []
    for r in rows:
        out = dict(r)
        for k in ("pantones", "image_names", "doc_names"):
            if out.get(k):
                try:
                    out[k] = json.loads(out[k])
                except (json.JSONDecodeError, TypeError):
                    out[k] = []
            else:
                out[k] = []
        records.append(_dev_record_to_xlsx(out))
    wb = _build_workbook(records, "Enquiries")
    _send_xlsx(handler, wb, "enquiries.xlsx")


def api_export_customers(handler):
    conn = db()
    companies = conn.execute(
        "SELECT * FROM companies ORDER BY id DESC").fetchall()
    records = []
    for c in companies:
        members = conn.execute(
            "SELECT * FROM members WHERE company_id = ? ORDER BY id",
            (c["id"],)).fetchall()
        ship_to = conn.execute(
            "SELECT * FROM ship_to WHERE company_id = ? ORDER BY is_default DESC, id",
            (c["id"],)).fetchall()
        projects = conn.execute(
            "SELECT * FROM projects WHERE company_id = ? ORDER BY id",
            (c["id"],)).fetchall()
        item = dict(c)
        item["members"] = [dict(m) for m in members]
        item["ship_to"] = [dict(s) for s in ship_to]
        item["projects"] = [dict(p) for p in projects]
        records.extend(_customer_record_to_xlsx(item))
    conn.close()
    wb = _build_workbook(records, "Customers")
    _send_xlsx(handler, wb, "customers.xlsx")


def dev_details_summary(d):
    parts = []
    if d.get("height") or d.get("width"):
        parts.append(f"{(d.get('height') or '?')} × {(d.get('width') or '?')} mm")
    if d.get("raised_height"):
        parts.append(f"raised {d.get('raised_height')} mm")
    # Split-color products (screen print / printed label / hang tag) store their
    # colors under `color_sides` (Front + Back); others use no_of_color/pantones.
    sides = d.get("color_sides")
    if sides and (sides.get("front") or sides.get("back")):
        for label in ("front", "back"):
            side = sides.get(label)
            if not side:
                continue
            n = side.get("no_of_color")
            if not n:
                continue
            cols = [(p.get("value") if isinstance(p, dict) else p)
                    for p in (side.get("pantones") or []) if p]
            try:
                plural = int(n) > 1
            except (TypeError, ValueError):
                plural = False
            parts.append(
                f"{label.capitalize()} {n} color{'s' if plural else ''}"
                + (f" ({', '.join(str(c) for c in cols)})" if cols else "")
            )
    elif d.get("no_of_color"):
        cols = [(p.get("value") if isinstance(p, dict) else p)
                for p in (d.get("pantones") or []) if p]
        n = d.get("no_of_color")
        try:
            plural = int(n) > 1
        except (TypeError, ValueError):
            plural = False
        parts.append(
            f"{n} color{'s' if plural else ''}"
            + (f" ({', '.join(str(c) for c in cols)})" if cols else "")
        )
    return " · ".join(parts)




# --- Sample images (for Development / Create Dummy + image pool) -------------

IMG_EXTS = (".png", ".jpg", ".jpeg", ".webp", ".gif")


def list_sample_images():
    """Scan SAMPLE_IMAGES_DIR once and return a {dir, files:[rel paths]} dict."""
    global _sample_images_cache
    if _sample_images_cache is not None:
        return _sample_images_cache
    files = []
    base = SAMPLE_IMAGES_DIR
    if os.path.isdir(base):
        for root, _dirs, names in os.walk(base):
            for n in names:
                if n.lower().endswith(IMG_EXTS):
                    full = os.path.join(root, n)
                    rel = os.path.relpath(full, base).replace(os.sep, "/")
                    files.append(rel)
    _sample_images_cache = {"dir": base, "files": files}
    return _sample_images_cache


def safe_sample_path(rel):
    """Resolve a relative image path under SAMPLE_IMAGES_DIR, guarding traversal."""
    base = os.path.abspath(SAMPLE_IMAGES_DIR)
    target = os.path.abspath(os.path.join(base, rel))
    if target != base and not target.startswith(base + os.sep):
        return None
    if not os.path.isfile(target):
        return None
    return target


def serve_sample_image(handler, rel):
    rel = unquote(rel)
    if "/../" in rel or rel.startswith("../") or rel.endswith("/.."):
        handler.send_response(404)
        handler.end_headers()
        return
    path = safe_sample_path(rel)
    if not path:
        handler.send_response(404)
        handler.end_headers()
        return
    ext = os.path.splitext(path)[1].lower()
    ctype = {
        ".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
        ".webp": "image/webp", ".gif": "image/gif",
    }.get(ext, "application/octet-stream")
    with open(path, "rb") as f:
        data = f.read()
    handler.send_response(200)
    handler.send_header("Content-Type", ctype)
    handler.send_header("Content-Length", str(len(data)))
    handler.send_header("Cache-Control", "public, max-age=3600")
    handler.end_headers()
    handler.wfile.write(data)


# --- Uploaded files (user-dropped images + documents) -----------------------

EXT_CTYPE = {
    ".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
    ".webp": "image/webp", ".gif": "image/gif",
    ".pdf": "application/pdf",
    ".txt": "text/plain; charset=utf-8",
    ".csv": "text/csv; charset=utf-8",
    ".doc": "application/msword",
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    ".xls": "application/vnd.ms-excel",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".zip": "application/zip",
}


def safe_upload_path(rel):
    """Resolve a relative path under UPLOADS_DIR, guarding traversal."""
    base = os.path.abspath(UPLOADS_DIR)
    target = os.path.abspath(os.path.join(base, rel))
    if target != base and not target.startswith(base + os.sep):
        return None
    if not os.path.isfile(target):
        return None
    return target


def resolve_upload_name(rel):
    """Map a stored doc name to an on-disk upload filename.

    New uploads are stored as 'uploads/<uuid>__<orig>' so an exact lookup works.
    Legacy records stored only the bare original (e.g. 'NXB30922CY074.pdf')
    which doesn't match the on-disk '<uuid>__<orig - extra>.pdf'. Best-effort
    match by exact basename or by leading-stem so those links still resolve.
    Returns the bare on-disk filename (relative to UPLOADS_DIR) or None.
    """
    if safe_upload_path(rel):
        return os.path.basename(rel)
    base = os.path.basename(rel)
    if not base:
        return None
    bstem = base.split("__", 1)[-1]  # drop uuid prefix if present
    try:
        entries = os.listdir(UPLOADS_DIR)
    except OSError:
        return None
    norm = lambda s: s.lower().replace(" ", "")
    # Drop a trailing extension from the query so "NXB30922CY074.pdf" matches an
    # on-disk original like "NXB30922CY074 - AW27_BAKER....pdf".
    q_stem = os.path.splitext(bstem)[0]
    q = norm(q_stem)
    if not q:
        return None
    # The meaningful leading token of the query, split on separators.
    q_token = re.split(r"[\s\-_.]", q_stem)[0]
    candidates = []
    for fn in entries:
        full = os.path.join(UPLOADS_DIR, fn)
        if not os.path.isfile(full):
            continue
        orig = os.path.splitext(fn.split("__", 1)[-1])[0]
        o = norm(orig)
        if o == q:
            candidates.append((0, fn))
        elif o.startswith(q):
            candidates.append((1, fn))
        elif q.startswith(o):
            candidates.append((2, fn))
        # Leading-token match: "NXB30922CY074" ⊂ "NXB30922CY074 - AW27…".
        elif q_token and o.startswith(q_token):
            candidates.append((3, fn))
    if not candidates:
        return None
    candidates.sort()
    return candidates[0][1]


def serve_upload(handler, rel):
    # The stored filename may contain spaces, "&", parentheses, etc. The HTTP
    # client percent-encodes these; decode back to the real on-disk name.
    rel = unquote(rel)
    # Reject any attempt to leave the uploads directory.
    if "/../" in rel or rel.startswith("../") or rel.endswith("/.."):
        handler.send_response(404)
        handler.end_headers()
        return
    path = safe_upload_path(rel)
    # Legacy bare doc names don't carry the <uuid>__ prefix; best-effort match.
    if not path:
        resolved = resolve_upload_name(rel)
        if resolved:
            path = os.path.join(UPLOADS_DIR, resolved)
    if not path:
        handler.send_response(404)
        handler.end_headers()
        return
    ext = os.path.splitext(path)[1].lower()
    ctype = EXT_CTYPE.get(ext, "application/octet-stream")
    with open(path, "rb") as f:
        data = f.read()
    handler.send_response(200)
    handler.send_header("Content-Type", ctype)
    handler.send_header("Content-Length", str(len(data)))
    handler.send_header("Cache-Control", "public, max-age=3600")
    # Force a download for documents (PDFs, etc.) so the browser saves them
    # instead of opening inline. Images stay inline so they render in <img>.
    if ctype.startswith("image/"):
        handler.send_header("Content-Disposition", "inline")
    else:
        # Show the original filename (drop the <uuid>__ prefix) on download.
        base = os.path.basename(path).split("__", 1)[-1]
        ascii_name = base.encode("ascii", "ignore").decode("ascii") or "download"
        handler.send_header(
            "Content-Disposition",
            "attachment; filename=\"%s\"; filename*=UTF-8''%s"
            % (ascii_name, quote(base)),
        )
    handler.end_headers()
    handler.wfile.write(data)


def _sanitize_filename(name):
    """Keep only the basename and strip path separators / parent dirs."""
    name = os.path.basename(name or "file")
    name = re.sub(r"[\/\\]", "_", name)
    name = name.replace("..", "_")
    return name or "file"


def api_upload_file(handler):
    """Accept multipart/form-data with a single 'file' field, persist it
    under data/uploads/<uuid>__<sanitized-original>, and return its path."""
    import cgi
    ctype = handler.headers.get("Content-Type", "")
    form = cgi.FieldStorage(
        fp=handler.rfile,
        headers=handler.headers,
        environ={"REQUEST_METHOD": "POST", "CONTENT_TYPE": ctype},
    )
    if "file" not in form:
        return json_response(handler, {"error": "no file field"}, 400)
    item = form["file"]
    if not hasattr(item, "filename") or not item.filename:
        return json_response(handler, {"error": "empty file"}, 400)
    orig = _sanitize_filename(item.filename)
    stored = uuid.uuid4().hex + "__" + orig
    with open(os.path.join(UPLOADS_DIR, stored), "wb") as f:
        f.write(item.file.read())
    rel = "uploads/" + stored          # served at GET /uploads/<rest>
    return json_response(handler, {"path": rel, "name": orig, "url": "/" + rel}, 201)


# --- API handlers ---------------------------------------------------------

def api_create_company(handler):
    data = read_json_body(handler)
    name = (data.get("name") or "").strip()
    suffix = (data.get("email_suffix") or "").strip().lstrip("@")
    if not name or not suffix:
        return json_response(handler, {"error": "name and email_suffix are required"}, 400)
    currency = (data.get("currency") or "").strip() or None
    payment_term = (data.get("payment_term") or "").strip() or None
    shipment_term = (data.get("shipment_term") or "").strip() or None
    conn = db()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO companies (name, email_suffix, currency, payment_term, "
        "shipment_term, created_at) VALUES (?, ?, ?, ?, ?, ?)",
        (name, suffix, currency, payment_term, shipment_term, now_iso()),
    )
    cid = cur.lastrowid
    conn.commit()
    conn.close()
    return json_response(
        handler,
        {"id": cid, "name": name, "email_suffix": suffix,
         "currency": currency, "payment_term": payment_term,
         "shipment_term": shipment_term},
        201,
    )


def api_list_companies(handler):
    conn = db()
    rows = conn.execute("SELECT * FROM companies ORDER BY id DESC").fetchall()
    conn.close()
    return json_response(handler, [dict(r) for r in rows])


def api_get_company(handler, cid):
    conn = db()
    comp = conn.execute("SELECT * FROM companies WHERE id = ?", (cid,)).fetchone()
    if not comp:
        conn.close()
        return json_response(handler, {"error": "not found"}, 404)
    members = conn.execute(
        "SELECT * FROM members WHERE company_id = ? ORDER BY id", (cid,)
    ).fetchall()
    ship_to = conn.execute(
        "SELECT * FROM ship_to WHERE company_id = ? ORDER BY is_default DESC, id", (cid,)
    ).fetchall()
    projects = conn.execute(
        "SELECT * FROM projects WHERE company_id = ? ORDER BY id", (cid,)
    ).fetchall()
    conn.close()
    out = dict(comp)
    out["members"] = [dict(m) for m in members]
    out["ship_to"] = [dict(s) for s in ship_to]
    out["projects"] = [dict(p) for p in projects]
    return json_response(handler, out)


def api_add_member(handler, cid):
    conn = db()
    comp = conn.execute("SELECT * FROM companies WHERE id = ?", (cid,)).fetchone()
    if not comp:
        conn.close()
        return json_response(handler, {"error": "company not found"}, 404)
    data = read_json_body(handler)
    name = (data.get("name") or "").strip()
    prefix = (data.get("email_prefix") or "").strip()
    title = (data.get("title") or "").strip()
    tel = (data.get("tel") or "").strip()
    if not (name and prefix and title and tel):
        conn.close()
        return json_response(handler, {"error": "name, email_prefix, title, tel are required"}, 400)
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO members (company_id, name, email_prefix, title, tel, created_at) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        (cid, name, prefix, title, tel, now_iso()),
    )
    mid = cur.lastrowid
    conn.commit()
    conn.close()
    return json_response(
        handler,
        {"id": mid, "company_id": cid, "name": name, "email_prefix": prefix,
         "title": title, "tel": tel},
        201,
    )


def api_update_company(handler, cid):
    conn = db()
    comp = conn.execute("SELECT * FROM companies WHERE id = ?", (cid,)).fetchone()
    if not comp:
        conn.close()
        return json_response(handler, {"error": "company not found"}, 404)
    data = read_json_body(handler)
    name = (data.get("name") or "").strip()
    suffix = (data.get("email_suffix") or "").strip().lstrip("@")
    if not name or not suffix:
        conn.close()
        return json_response(handler, {"error": "name and email_suffix are required"}, 400)
    currency = (data.get("currency") or "").strip() or None
    payment_term = (data.get("payment_term") or "").strip() or None
    shipment_term = (data.get("shipment_term") or "").strip() or None
    conn.execute(
        "UPDATE companies SET name = ?, email_suffix = ?, currency = ?, "
        "payment_term = ?, shipment_term = ? WHERE id = ?",
        (name, suffix, currency, payment_term, shipment_term, cid),
    )
    conn.commit()
    conn.close()
    return json_response(
        handler,
        {"id": cid, "name": name, "email_suffix": suffix,
         "currency": currency, "payment_term": payment_term,
         "shipment_term": shipment_term},
        200,
    )


def api_update_member(handler, mid):
    conn = db()
    mem = conn.execute("SELECT * FROM members WHERE id = ?", (mid,)).fetchone()
    if not mem:
        conn.close()
        return json_response(handler, {"error": "member not found"}, 404)
    data = read_json_body(handler)
    name = (data.get("name") or "").strip()
    prefix = (data.get("email_prefix") or "").strip()
    title = (data.get("title") or "").strip()
    tel = (data.get("tel") or "").strip()
    if not (name and prefix and title and tel):
        conn.close()
        return json_response(handler, {"error": "name, email_prefix, title, tel are required"}, 400)
    conn.execute(
        "UPDATE members SET name = ?, email_prefix = ?, title = ?, tel = ? WHERE id = ?",
        (name, prefix, title, tel, mid),
    )
    conn.commit()
    conn.close()
    return json_response(
        handler,
        {"id": mid, "name": name, "email_prefix": prefix, "title": title, "tel": tel},
        200,
    )


def api_delete_company(handler, cid):
    conn = db()
    comp = conn.execute("SELECT * FROM companies WHERE id = ?", (cid,)).fetchone()
    if not comp:
        conn.close()
        return json_response(handler, {"error": "company not found"}, 404)
    # delete members, ship-to addresses, projects, and developments first
    # (FK enforcement may be off in older SQLite sessions, and developments has
    # no FK at all), then the company row itself. This removes the whole company
    # and every record tied to it.
    conn.execute("DELETE FROM members WHERE company_id = ?", (cid,))
    conn.execute("DELETE FROM ship_to WHERE company_id = ?", (cid,))
    conn.execute("DELETE FROM projects WHERE company_id = ?", (cid,))
    conn.execute("DELETE FROM developments WHERE company_id = ?", (cid,))
    conn.execute("DELETE FROM enquiries WHERE company_id = ?", (cid,))
    conn.execute("DELETE FROM companies WHERE id = ?", (cid,))
    conn.commit()
    conn.close()
    return json_response(handler, {"ok": True, "id": cid}, 200)


def api_delete_member(handler, mid):
    conn = db()
    mem = conn.execute("SELECT * FROM members WHERE id = ?", (mid,)).fetchone()
    if not mem:
        conn.close()
        return json_response(handler, {"error": "member not found"}, 404)
    conn.execute("DELETE FROM members WHERE id = ?", (mid,))
    conn.commit()
    conn.close()
    return json_response(handler, {"ok": True, "id": mid}, 200)


def api_list_customers(handler):
    """Flat join for the View page: one row per (company) with members nested."""
    conn = db()
    companies = conn.execute("SELECT * FROM companies ORDER BY id DESC").fetchall()
    out = []
    for c in companies:
        members = conn.execute(
            "SELECT * FROM members WHERE company_id = ? ORDER BY id", (c["id"],)
        ).fetchall()
        ship_to = conn.execute(
            "SELECT * FROM ship_to WHERE company_id = ? ORDER BY is_default DESC, id", (c["id"],)
        ).fetchall()
        projects = conn.execute(
            "SELECT * FROM projects WHERE company_id = ? ORDER BY id", (c["id"],)
        ).fetchall()
        item = dict(c)
        item["members"] = [dict(m) for m in members]
        item["ship_to"] = [dict(s) for s in ship_to]
        item["projects"] = [dict(p) for p in projects]
        out.append(item)
    conn.close()
    return json_response(handler, out)


# --- Ship-To handlers -------------------------------------------------------

def api_list_ship_to(handler, cid):
    conn = db()
    rows = conn.execute(
        "SELECT * FROM ship_to WHERE company_id = ? ORDER BY is_default DESC, id", (cid,)
    ).fetchall()
    conn.close()
    return json_response(handler, [dict(r) for r in rows])


def api_add_ship_to(handler, cid):
    conn = db()
    comp = conn.execute("SELECT * FROM companies WHERE id = ?", (cid,)).fetchone()
    if not comp:
        conn.close()
        return json_response(handler, {"error": "company not found"}, 404)
    data = read_json_body(handler)
    address = (data.get("address") or "").strip()
    if not address:
        conn.close()
        return json_response(handler, {"error": "address is required"}, 400)
    # First address for a company becomes the default automatically.
    existing = conn.execute(
        "SELECT COUNT(*) AS n FROM ship_to WHERE company_id = ?", (cid,)
    ).fetchone()["n"]
    is_default = 1 if existing == 0 else 0
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO ship_to (company_id, address, is_default, created_at) "
        "VALUES (?, ?, ?, ?)",
        (cid, address, is_default, now_iso()),
    )
    sid = cur.lastrowid
    conn.commit()
    conn.close()
    return json_response(
        handler,
        {"id": sid, "company_id": cid, "address": address, "is_default": is_default},
        201,
    )


def api_set_default_ship_to(handler, sid):
    conn = db()
    row = conn.execute("SELECT * FROM ship_to WHERE id = ?", (sid,)).fetchone()
    if not row:
        conn.close()
        return json_response(handler, {"error": "not found"}, 404)
    cid = row["company_id"]
    # Clear all defaults for the company, then mark this one.
    conn.execute("UPDATE ship_to SET is_default = 0 WHERE company_id = ?", (cid,))
    conn.execute("UPDATE ship_to SET is_default = 1 WHERE id = ?", (sid,))
    conn.commit()
    conn.close()
    return json_response(handler, {"ok": True, "id": sid}, 200)


def api_delete_ship_to(handler, sid):
    conn = db()
    row = conn.execute("SELECT * FROM ship_to WHERE id = ?", (sid,)).fetchone()
    if not row:
        conn.close()
        return json_response(handler, {"error": "not found"}, 404)
    was_default = row["is_default"]
    cid = row["company_id"]
    conn.execute("DELETE FROM ship_to WHERE id = ?", (sid,))
    # If we removed the default and others remain, promote the first remaining.
    if was_default:
        first = conn.execute(
            "SELECT * FROM ship_to WHERE company_id = ? ORDER BY id LIMIT 1", (cid,)
        ).fetchone()
        if first:
            conn.execute("UPDATE ship_to SET is_default = 1 WHERE id = ?", (first["id"],))
    conn.commit()
    conn.close()
    return json_response(handler, {"ok": True, "id": sid}, 200)


# --- Project handlers -------------------------------------------------------

def api_list_projects(handler, cid):
    conn = db()
    rows = conn.execute(
        "SELECT * FROM projects WHERE company_id = ? ORDER BY id", (cid,)
    ).fetchall()
    conn.close()
    return json_response(handler, [dict(r) for r in rows])


def api_add_project(handler, cid):
    conn = db()
    comp = conn.execute("SELECT * FROM companies WHERE id = ?", (cid,)).fetchone()
    if not comp:
        conn.close()
        return json_response(handler, {"error": "company not found"}, 404)
    data = read_json_body(handler)
    name = (data.get("name") or "").strip()
    if not name:
        conn.close()
        return json_response(handler, {"error": "name is required"}, 400)
    try:
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO projects (company_id, name, created_at) VALUES (?, ?, ?)",
            (cid, name, now_iso()),
        )
        pid = cur.lastrowid
        conn.commit()
        conn.close()
        return json_response(handler, {"id": pid, "company_id": cid, "name": name}, 201)
    except sqlite3.IntegrityError:
        conn.close()
        return json_response(
            handler, {"error": "a project with this name already exists for this company"}, 409
        )


def api_delete_project(handler, pid):
    conn = db()
    row = conn.execute("SELECT * FROM projects WHERE id = ?", (pid,)).fetchone()
    if not row:
        conn.close()
        return json_response(handler, {"error": "not found"}, 404)
    conn.execute("DELETE FROM projects WHERE id = ?", (pid,))
    conn.commit()
    conn.close()
    return json_response(handler, {"ok": True, "id": pid}, 200)


# --- Development handlers --------------------------------------------------

def _dev_row_to_payload(row):
    out = dict(row)
    for k in ("pantones", "image_names", "doc_names"):
        if out.get(k):
            try:
                out[k] = json.loads(out[k])
            except (json.JSONDecodeError, TypeError):
                out[k] = []
        else:
            out[k] = []
    for k in ("material", "special", "remake"):
        if out.get(k):
            try:
                out[k] = json.loads(out[k])
            except (json.JSONDecodeError, TypeError):
                out[k] = out[k]
        else:
            out[k] = None
    if out.get("color_sides"):
        try:
            out["color_sides"] = json.loads(out["color_sides"])
        except (json.JSONDecodeError, TypeError):
            out["color_sides"] = None
    else:
        out["color_sides"] = None
    return out


def api_list_developments(handler):
    conn = db()
    rows = conn.execute("SELECT * FROM developments ORDER BY id DESC").fetchall()
    conn.close()
    return json_response(handler, [_dev_row_to_payload(r) for r in rows])


def api_get_development(handler, did):
    conn = db()
    row = conn.execute("SELECT * FROM developments WHERE id = ?", (did,)).fetchone()
    conn.close()
    if not row:
        return json_response(handler, {"error": "not found"}, 404)
    return json_response(handler, _dev_row_to_payload(row))


def _dev_validate(data):
    return (data.get("company_name") or "").strip() and \
           (data.get("item_name") or "").strip() and \
           (data.get("product_type") or "").strip()


def _dev_insert_or_update(conn, did, data):
    company_name = (data.get("company_name") or "").strip()
    item_name = (data.get("item_name") or "").strip()
    product_type = (data.get("product_type") or "").strip()
    pantones = data.get("pantones")
    image_names = data.get("image_names")
    if isinstance(pantones, (list, dict)):
        pantones = json.dumps(pantones, ensure_ascii=False)
    if isinstance(image_names, (list, dict)):
        image_names = json.dumps(image_names, ensure_ascii=False)
    doc_names = data.get("doc_names")
    if isinstance(doc_names, (list, dict)):
        doc_names = json.dumps(doc_names, ensure_ascii=False)
    material = data.get("material")
    special = data.get("special")
    remake = data.get("remake")
    if isinstance(material, (list, dict)):
        material = json.dumps(material, ensure_ascii=False)
    if isinstance(special, (list, dict)):
        special = json.dumps(special, ensure_ascii=False)
    if isinstance(remake, (list, dict)):
        remake = json.dumps(remake, ensure_ascii=False)
    color_sides = data.get("color_sides")
    if isinstance(color_sides, (list, dict)):
        color_sides = json.dumps(color_sides, ensure_ascii=False)
    # A new development starts with status "Created" unless the payload says
    # otherwise. Updates never touch status (it is driven by Follow Ups).
    status_val = (data.get("status") or "").strip() or "Created"
    vals = (
        data.get("company_id") if did is None else (data.get("company_id") if data.get("company_id") is not None else None),
        company_name,
        data.get("member_id") if data.get("member_id") is not None else None,
        (data.get("member_name") or "").strip() or None,
        data.get("project_id") if data.get("project_id") is not None else None,
        (data.get("project_name") or "").strip() or None,
        item_name,
        product_type,
        _to_float(data.get("height")),
        _to_float(data.get("width")),
        _to_float(data.get("raised_height")),
        _to_int(data.get("no_of_color")),
        pantones,
        image_names,
        doc_names,
        material,
        special,
        remake,
        color_sides,
    )
    if did is None:
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO developments "
            "(company_id, company_name, member_id, member_name, project_id, project_name, "
            "item_name, product_type, height, width, raised_height, no_of_color, pantones, "
            "image_names, doc_names, material, special, remake, color_sides, status, created_at, updated_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            vals + (status_val, now_iso(), now_iso()),
        )
        return cur.lastrowid
    conn.execute(
        "UPDATE developments SET "
        "company_id=?, company_name=?, member_id=?, member_name=?, project_id=?, project_name=?, "
        "item_name=?, product_type=?, height=?, width=?, raised_height=?, no_of_color=?, "
        "pantones=?, image_names=?, doc_names=?, material=?, special=?, remake=?, color_sides=?, updated_at=? WHERE id=?",
        vals + (now_iso(), did),
    )
    return did


def _to_float(v):
    try:
        if v in (None, ""):
            return None
        return float(v)
    except (TypeError, ValueError):
        return None


def _to_int(v):
    try:
        if v in (None, ""):
            return None
        return int(v)
    except (TypeError, ValueError):
        return None


def api_create_development(handler):
    data = read_json_body(handler)
    if not _dev_validate(data):
        return json_response(handler, {"error": "company_name, item_name, product_type are required"}, 400)
    conn = db()
    did = _dev_insert_or_update(conn, None, data)
    conn.commit()
    conn.close()
    return json_response(handler, {"id": did}, 201)


def api_update_development(handler, did):
    conn = db()
    row = conn.execute("SELECT * FROM developments WHERE id = ?", (did,)).fetchone()
    if not row:
        conn.close()
        return json_response(handler, {"error": "not found"}, 404)
    data = read_json_body(handler)
    if not _dev_validate(data):
        conn.close()
        return json_response(handler, {"error": "company_name, item_name, product_type are required"}, 400)
    _dev_insert_or_update(conn, did, data)
    conn.commit()
    conn.close()
    return json_response(handler, {"id": did}, 200)


def api_delete_development(handler, did):
    conn = db()
    row = conn.execute("SELECT * FROM developments WHERE id = ?", (did,)).fetchone()
    if not row:
        conn.close()
        return json_response(handler, {"error": "not found"}, 404)
    conn.execute("DELETE FROM developments WHERE id = ?", (did,))
    conn.commit()
    conn.close()
    return json_response(handler, {"ok": True, "id": did}, 200)


# --- Development follow-up handlers -------------------------------------

def _followup_row_to_payload(row):
    out = dict(row)
    for k in ("image_names", "doc_names"):
        if out.get(k):
            try:
                out[k] = json.loads(out[k])
            except (json.JSONDecodeError, TypeError):
                out[k] = []
        else:
            out[k] = []
    return out


def _json_array(v):
    """Normalise a submitted list (or list-shaped string) to a JSON array."""
    if isinstance(v, str):
        try:
            parsed = json.loads(v)
        except (json.JSONDecodeError, TypeError):
            parsed = v
        if isinstance(parsed, list):
            v = parsed
        else:
            v = [v] if parsed else []
    if not isinstance(v, (list, tuple)):
        v = []
    return json.dumps([str(x) for x in v], ensure_ascii=False)


def api_create_followup(handler, did):
    conn = db()
    dev = conn.execute("SELECT id FROM developments WHERE id = ?", (did,)).fetchone()
    if not dev:
        conn.close()
        return json_response(handler, {"error": "not found"}, 404)
    data = read_json_body(handler)
    category = (data.get("category") or "").strip()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO followups (development_id, category, note, image_names, doc_names, created_at) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        (
            did,
            category or None,
            data.get("note") or None,
            _json_array(data.get("image_names")),
            _json_array(data.get("doc_names")),
            now_iso(),
        ),
    )
    fid = cur.lastrowid
    if category:
        conn.execute(
            "UPDATE developments SET status = ?, updated_at = ? WHERE id = ?",
            (category, now_iso(), did),
        )
    row = conn.execute("SELECT * FROM followups WHERE id = ?", (fid,)).fetchone()
    conn.commit()
    conn.close()
    return json_response(handler, _followup_row_to_payload(row), 201)


def api_update_followup(handler, did, fid):
    conn = db()
    dev = conn.execute("SELECT id FROM developments WHERE id = ?", (did,)).fetchone()
    if not dev:
        conn.close()
        return json_response(handler, {"error": "not found"}, 404)
    row = conn.execute(
        "SELECT * FROM followups WHERE id = ? AND development_id = ?", (fid, did)
    ).fetchone()
    if not row:
        conn.close()
        return json_response(handler, {"error": "not found"}, 404)
    data = read_json_body(handler)
    category = (data.get("category") or "").strip()
    old_category = (row["category"] or "").strip() if row["category"] else ""
    conn.execute(
        "UPDATE followups SET category = ?, note = ?, image_names = ?, doc_names = ? WHERE id = ?",
        (
            category or None,
            data.get("note") or None,
            _json_array(data.get("image_names")),
            _json_array(data.get("doc_names")),
            fid,
        ),
    )
    # The development's Status follows the latest non-empty Category.
    if category and category != old_category:
        conn.execute(
            "UPDATE developments SET status = ?, updated_at = ? WHERE id = ?",
            (category, now_iso(), did),
        )
    updated = conn.execute("SELECT * FROM followups WHERE id = ?", (fid,)).fetchone()
    conn.commit()
    conn.close()
    return json_response(handler, _followup_row_to_payload(updated), 200)


def api_list_followups(handler, did):
    conn = db()
    dev = conn.execute("SELECT id FROM developments WHERE id = ?", (did,)).fetchone()
    if not dev:
        conn.close()
        return json_response(handler, {"error": "not found"}, 404)
    rows = conn.execute(
        "SELECT * FROM followups WHERE development_id = ? ORDER BY id DESC", (did,)
    ).fetchall()
    conn.close()
    return json_response(handler, [_followup_row_to_payload(r) for r in rows])


# --- Enquiry handlers ----------------------------------------------------

def _enquiry_row_to_payload(row):
    out = dict(row)
    for k in ("pantones", "image_names", "doc_names"):
        if out.get(k):
            try:
                out[k] = json.loads(out[k])
            except (json.JSONDecodeError, TypeError):
                out[k] = []
        else:
            out[k] = []
    for k in ("material", "special", "remake"):
        if out.get(k):
            try:
                out[k] = json.loads(out[k])
            except (json.JSONDecodeError, TypeError):
                out[k] = out[k]
        else:
            out[k] = None
    if out.get("color_sides"):
        try:
            out["color_sides"] = json.loads(out["color_sides"])
        except (json.JSONDecodeError, TypeError):
            out["color_sides"] = None
    else:
        out["color_sides"] = None
    return out


def api_list_enquiries(handler):
    conn = db()
    rows = conn.execute("SELECT * FROM enquiries ORDER BY id DESC").fetchall()
    conn.close()
    return json_response(handler, [_enquiry_row_to_payload(r) for r in rows])


def api_get_enquiry(handler, eid):
    conn = db()
    row = conn.execute("SELECT * FROM enquiries WHERE id = ?", (eid,)).fetchone()
    conn.close()
    if not row:
        return json_response(handler, {"error": "not found"}, 404)
    return json_response(handler, _enquiry_row_to_payload(row))


def _enquiry_validate(data):
    return bool((data.get("company_name") or "").strip())


def _enquiry_insert_or_update(conn, eid, data):
    company_name = (data.get("company_name") or "").strip()
    item_name = (data.get("item_name") or "").strip()
    product_type = (data.get("product_type") or "").strip()
    pantones = data.get("pantones")
    if isinstance(pantones, (list, dict)):
        pantones = json.dumps(pantones, ensure_ascii=False)
    image_names = data.get("image_names")
    if isinstance(image_names, (list, dict)):
        image_names = json.dumps(image_names, ensure_ascii=False)
    doc_names = data.get("doc_names")
    if isinstance(doc_names, (list, dict)):
        doc_names = json.dumps(doc_names, ensure_ascii=False)
    color_sides = data.get("color_sides")
    if isinstance(color_sides, (list, dict)):
        color_sides = json.dumps(color_sides, ensure_ascii=False)
    notes = (data.get("notes") or "").strip() or None
    vals = (
        data.get("company_id") if data.get("company_id") is not None else None,
        company_name,
        data.get("member_id") if data.get("member_id") is not None else None,
        (data.get("member_name") or "").strip() or None,
        data.get("project_id") if data.get("project_id") is not None else None,
        (data.get("project_name") or "").strip() or None,
        item_name,
        product_type,
        _to_float(data.get("height")),
        _to_float(data.get("width")),
        _to_float(data.get("raised_height")),
        _to_int(data.get("no_of_color")),
        pantones,
        image_names,
        doc_names,
        color_sides,
        notes,
    )
    if eid is None:
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO enquiries "
            "(company_id, company_name, member_id, member_name, project_id, project_name, "
            "item_name, product_type, height, width, raised_height, no_of_color, pantones, "
            "image_names, doc_names, color_sides, notes, created_at, updated_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            vals + (now_iso(), now_iso()),
        )
        return cur.lastrowid
    conn.execute(
        "UPDATE enquiries SET "
        "company_id=?, company_name=?, member_id=?, member_name=?, project_id=?, project_name=?, "
        "item_name=?, product_type=?, height=?, width=?, raised_height=?, no_of_color=?, "
        "pantones=?, image_names=?, doc_names=?, color_sides=?, notes=?, updated_at=? WHERE id=?",
        vals + (now_iso(), eid),
    )
    return eid


def api_create_enquiry(handler):
    data = read_json_body(handler)
    if not _enquiry_validate(data):
        return json_response(handler, {"error": "company_name is required"}, 400)
    conn = db()
    eid = _enquiry_insert_or_update(conn, None, data)
    conn.commit()
    conn.close()
    return json_response(handler, {"id": eid}, 201)


def api_update_enquiry(handler, eid):
    conn = db()
    row = conn.execute("SELECT * FROM enquiries WHERE id = ?", (eid,)).fetchone()
    if not row:
        conn.close()
        return json_response(handler, {"error": "not found"}, 404)
    data = read_json_body(handler)
    if not _enquiry_validate(data):
        conn.close()
        return json_response(handler, {"error": "company_name is required"}, 400)
    _enquiry_insert_or_update(conn, eid, data)
    conn.commit()
    conn.close()
    return json_response(handler, {"id": eid}, 200)


def api_delete_enquiry(handler, eid):
    conn = db()
    row = conn.execute("SELECT * FROM enquiries WHERE id = ?", (eid,)).fetchone()
    if not row:
        conn.close()
        return json_response(handler, {"error": "not found"}, 404)
    conn.execute("DELETE FROM enquiries WHERE id = ?", (eid,))
    conn.commit()
    conn.close()
    return json_response(handler, {"ok": True, "id": eid}, 200)


# --- Router ---------------------------------------------------------------

class Handler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=STATIC_DIR, **kwargs)

    def end_headers(self):
        # Never let the browser cache static assets. app.js / css are edited
        # frequently during development; a cached old copy silently hides
        # every fix. Force re-fetch on every request.
        self.send_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
        self.send_header("Pragma", "no-cache")
        self.send_header("Expires", "0")
        super().end_headers()

    def _serve_sample_list(self):
        return json_response(self, list_sample_images())

    def _route(self):
        parsed = urlparse(self.path)
        path = parsed.path
        method = self.command

        if method == "OPTIONS":
            self.send_response(204)
            self.send_header("Access-Control-Allow-Origin", "*")
            self.send_header("Access-Control-Allow-Methods", "GET, POST, PUT, DELETE, OPTIONS")
            self.send_header("Access-Control-Allow-Headers", "Content-Type")
            self.end_headers()
            return True

        if path == "/api/companies" and method == "POST":
            api_create_company(self); return True
        if path == "/api/companies" and method == "GET":
            api_list_companies(self); return True
        if path == "/api/customers" and method == "GET":
            api_list_customers(self); return True
        if path == "/api/export/developments" and method == "GET":
            api_export_developments(self); return True
        if path == "/api/export/enquiries" and method == "GET":
            api_export_enquiries(self); return True
        if path == "/api/export/customers" and method == "GET":
            api_export_customers(self); return True
        if path == "/api/sample-images" and method == "GET":
            return self._serve_sample_list()
        if path == "/api/uploads" and method == "POST":
            api_upload_file(self); return True
        if path == "/api/developments" and method == "POST":
            api_create_development(self); return True
        if path == "/api/developments" and method == "GET":
            api_list_developments(self); return True
        if path.startswith("/sample-images/"):
            serve_sample_image(self, path[len("/sample-images/"):]); return True
        if path.startswith("/uploads/"):
            serve_upload(self, path[len("/uploads/"):]); return True
        if path.startswith("/api/members/"):
            rest = path[len("/api/members/"):]
            if rest.isdigit():
                mid = int(rest)
                if method == "PUT":
                    api_update_member(self, mid); return True
                if method == "DELETE":
                    api_delete_member(self, mid); return True
            return False

        if path.startswith("/api/ship-to/"):
            rest = path[len("/api/ship-to/"):]
            parts = rest.split("/")
            if len(parts) == 1 and parts[0].isdigit():
                cid = int(parts[0])
                if method == "GET":
                    api_list_ship_to(self, cid); return True
                if method == "POST":
                    api_add_ship_to(self, cid); return True
            elif len(parts) == 2 and parts[0].isdigit():
                sid = int(parts[0])
                if parts[1] == "default" and method == "PUT":
                    api_set_default_ship_to(self, sid); return True
                if method == "DELETE":
                    api_delete_ship_to(self, sid); return True
            return False

        if path.startswith("/api/projects/"):
            rest = path[len("/api/projects/"):]
            parts = rest.split("/")
            if len(parts) == 1 and parts[0].isdigit():
                cid = int(parts[0])
                if method == "GET":
                    api_list_projects(self, cid); return True
                if method == "POST":
                    api_add_project(self, cid); return True
            elif len(parts) == 2 and parts[0].isdigit() and method == "DELETE":
                pid = int(parts[0])
                api_delete_project(self, pid); return True
            return False

        if path.startswith("/api/companies/"):
            rest = path[len("/api/companies/"):]
            if rest.isdigit():
                cid = int(rest)
                if method == "GET":
                    api_get_company(self, cid); return True
                if method == "PUT":
                    api_update_company(self, cid); return True
                if method == "DELETE":
                    api_delete_company(self, cid); return True
            else:
                # /api/companies/<id>/members
                parts = rest.split("/")
                if len(parts) == 2 and parts[0].isdigit() and parts[1] == "members":
                    if method == "POST":
                        api_add_member(self, int(parts[0])); return True

        if path.startswith("/api/developments/"):
            rest = path[len("/api/developments/"):]
            parts = rest.split("/")
            if len(parts) == 3 and parts[0].isdigit() and parts[1] == "followups" and parts[2].isdigit():
                did = int(parts[0])
                fid = int(parts[2])
                if method == "PUT":
                    api_update_followup(self, did, fid); return True
                return True
            if len(parts) == 2 and parts[0].isdigit() and parts[1] == "followups":
                did = int(parts[0])
                if method == "GET":
                    api_list_followups(self, did); return True
                if method == "POST":
                    api_create_followup(self, did); return True
                return True
            if rest.isdigit():
                did = int(rest)
                if method == "GET":
                    api_get_development(self, did); return True
                if method == "PUT":
                    api_update_development(self, did); return True
                if method == "DELETE":
                    api_delete_development(self, did); return True

        if path == "/api/enquiries" and method == "POST":
            api_create_enquiry(self); return True
        if path == "/api/enquiries" and method == "GET":
            api_list_enquiries(self); return True
        if path == "/api/options" and method == "GET":
            api_get_options(self); return True
        if path == "/api/options" and method == "POST":
            api_add_option(self); return True
        if path == "/api/options/groups" and method == "GET":
            api_list_groups(self); return True
        if path == "/api/options/groups" and method == "POST":
            api_register_group(self); return True
        if path == "/api/options/reorder" and method == "PUT":
            api_reorder_options(self); return True
        if path.startswith("/api/options/"):
            rest = path[len("/api/options/"):]
            if rest.isdigit():
                oid = int(rest)
                if method == "PUT":
                    api_rename_option(self, oid); return True
                if method == "DELETE":
                    api_delete_option(self, oid); return True
            return False
        if path == "/api/product-type-factory" and method == "GET":
            api_get_product_type_factory(self); return True
        if path == "/api/product-type-factory" and method == "POST":
            api_add_ptf(self); return True
        if path == "/api/product-type-factory/list" and method == "POST":
            api_add_ptf_list(self); return True
        if path == "/api/product-type-factory/list" and method == "PUT":
            api_rename_ptf_list(self); return True
        if path == "/api/product-type-factory/list" and method == "DELETE":
            api_delete_ptf_list(self); return True
        if path == "/api/product-type-factory/list/type" and method == "POST":
            api_set_ptf_list_input_type(self); return True
        if path == "/api/product-type-factory/reorder" and method == "PUT":
            api_reorder_ptf(self); return True
        if path == "/api/product-type-factory/reorder-lists" and method == "PUT":
            api_reorder_ptf_lists(self); return True
        if path.startswith("/api/product-type-factory/"):
            rest = path[len("/api/product-type-factory/"):]
            if rest.isdigit():
                oid = int(rest)
                if method == "PUT":
                    api_rename_ptf(self, oid); return True
                if method == "DELETE":
                    api_delete_ptf(self, oid); return True
            return False

        if path.startswith("/api/enquiries/"):
            rest = path[len("/api/enquiries/"):]
            if rest.isdigit():
                eid = int(rest)
                if method == "GET":
                    api_get_enquiry(self, eid); return True
                if method == "PUT":
                    api_update_enquiry(self, eid); return True
                if method == "DELETE":
                    api_delete_enquiry(self, eid); return True

        return False

    def do_GET(self):
        if self._route():
            return
        super().do_GET()

    def do_POST(self):
        if self._route():
            return
        super().do_POST()

    def do_PUT(self):
        if self._route():
            return
        super().do_PUT()

    def do_DELETE(self):
        if self._route():
            return
        super().do_DELETE()

    def do_OPTIONS(self):
        if self._route():
            return
        super().do_OPTIONS()

    def log_message(self, fmt, *args):
        pass  # quiet


def main():
    init_db()
    server = HTTPServer(("0.0.0.0", PORT), Handler)
    print(f"FC server on http://localhost:{PORT}  (LAN: http://{os.environ.get('COMPUTERNAME','localhost')}:{PORT})")
    print(f"DB: {DB_PATH}")
    print("CTRL+C to stop.")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nStopped.")
        server.server_close()


if __name__ == "__main__":
    main()
